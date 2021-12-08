# basic
import os
import re
import time
import numpy as np 
import pandas as pd 

# excel 
from openpyxl import load_workbook
from pandas import ExcelWriter 
from openpyxl import Workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas.io import excel


class Excel_Work():
    """
    寫好的 表格 excle 輸出 , 不能含特殊字元 like : *  , 直接 input dataframe
    """
    def write_excel(self,df,line,tabel_style,tabel_style_name,excel_sheet_name):       
        
        wb = Workbook()

        ws = wb.active
        ws.title = str(excel_sheet_name)

        ws.append(list(df.columns))
        for i in range(df.shape[0]):
            row=df.iloc[i].values
            row = list(row)
            ws.append(row) # dataframe dataset 

        row=df.shape[0]+1  # df row     -->  data shape
        column=df.shape[1] # df column  -->  data shape
        column=chr(64+column)
        
        tab = Table(displayName="Table",ref="A1:"+str(column)+str(row))
        # Add a default style with striped rows and banded columns
        if tabel_style == True:
            style = TableStyleInfo(name=str(tabel_style_name), showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        else:
            style = TableStyleInfo(showFirstColumn=True,showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        #---------------------------------------設定邊框--------------------------------------------
        if line == True :
            rows = ws.max_row
            cols = ws.max_column
            # 線條 type 設定
            font2 = Font(name='微软雅黑', size=11)
            line_t = Side(style='thin', color='000000') 
            line_m = Side(style='medium', color='000000')
            # 設定表格內置中
            alignment = Alignment(horizontal='center', vertical='center') 
            # 邊框線條設定
            border1 = Border(top=line_m, bottom=line_t, left=line_t, right=line_t) 
            border2 = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)
            # 字形 style
            sty1 = NamedStyle(name='sty1', font=font2, 
                    border=border1, alignment=alignment)
            sty2 = NamedStyle(name='sty2', font=font2, border=border2, alignment=alignment)
            # 寫入 , 並帶設定字形
            for r in range(1, rows+1):
                for c in range(1, cols+1):
                    if r == 2:
                        ws.cell(r, c).style = sty1 # 第一欄位 columns 字形設定粗體 !
                    else:
                        ws.cell(r, c).style = sty2
            #-------------------------------------反白------------------------------------------------
        return wb
    def append_excel(self,excel_path,df,excel_sheet_name):
        book = load_workbook(excel_path)
        writer = ExcelWriter(excel_path, engine='openpyxl')  # pylint: disable=abstract-class-instantiated
        writer.book = book
        df.to_excel(writer,str(excel_sheet_name),index=False)
        writer.save()

class Address_DF():


    def Top_20_DF(self,df):

        def Address_Accounting_Numbers(number):

            try:

                number = str(number).replace(",","")
                number = float(number)
                return number

            except:
                
                if number == "" :
                    return 0 
                else :
                    return number

        # df["淨申贖金額"] = df.apply(lambda x : Address_Accounting_Numbers(x["淨申贖金額"]),axis=1)
        # df["申購金額"]   = df.apply(lambda x : Address_Accounting_Numbers(x["申購金額"]),axis=1)
        # df["買回金額"]   = df.apply(lambda x : Address_Accounting_Numbers(x["買回金額"]),axis=1)

        申購金額_df        = df.sort_values(by="申購金額",ascending=False).reset_index(drop=True)
        買回金額_df        = df.sort_values(by="買回金額",ascending=False).reset_index(drop=True)
        淨申贖金額_df      = df.sort_values(by="淨申贖金額",ascending=False).reset_index(drop=True)

        申購金額_result    = 申購金額_df[申購金額_df.index < 20]
        買回金額_result    = 買回金額_df[申購金額_df.index < 20]
        淨申贖金額_result  = 淨申贖金額_df[淨申贖金額_df.index < 20]

        return 申購金額_result ,買回金額_result ,淨申贖金額_result

    def 法人級別種類(self,法人級別_df) :

        法人級別_output = pd.DataFrame()
        法人級別_filter_dic = {
            "ALL"      : ["I","IT"],
            "聯博"     : ["IA","I","IT"],
            "PIMCO"    : ["H","I","IT"],
            "富達"     : ["Y","I","IT"], 
        }


        for i in range(法人級別_df.shape[0]):
            fund_name = 法人級別_df['基金級別名稱'][i]


            if "PIMCO" in fund_name : # 判別基金為 PIMCO　的
                index = fund_name.index("PIMCO")
                fund_condition_name = fund_name[index+5:]
                for PICOM_法人級別 in 法人級別_filter_dic['PIMCO'] : # PIMCO-法人級別的判別　Words 
                    if PICOM_法人級別 in fund_condition_name :
                        法人級別_output = 法人級別_output.append(法人級別_df.iloc[i],ignore_index=True)
            
            elif "富達" in fund_name : # 判別基金為 富達　的
                index = fund_name.index("富達")
                fund_condition_name = fund_name[index+2:]
                for 富達_法人級別 in 法人級別_filter_dic['富達'] :  # 富達-法人級別的判別　Words 
                    if 富達_法人級別 in fund_condition_name :
                        法人級別_output = 法人級別_output.append(法人級別_df.iloc[i],ignore_index=True)

            
            elif "聯博" in fund_name : # 判別基金為 聯博　的
                index = fund_name.index("聯博")
                fund_condition_name = fund_name[index+2:]
                for 聯博_法人級別 in 法人級別_filter_dic['聯博'] : # 聯博-法人級別的判別　Words 
                    if 聯博_法人級別 in fund_condition_name :
                        法人級別_output = 法人級別_output.append(法人級別_df.iloc[i],ignore_index=True)

            else : 
                for 法人級別 in 法人級別_filter_dic['ALL'] : 
                    
                    # 解決 AI 名稱被判定為 法人機構問題
                    try:
                        fund_name = fund_name.replace("AI","")
                    except:
                        pass

                    if 法人級別 in fund_name  :
                        法人級別_output = 法人級別_output.append(法人級別_df.iloc[i],ignore_index=True)

        法人級別_output = 法人級別_output[["基金級別名稱","基金類型","申購金額","買回金額","淨申購金額","國人持有基金","基金規模","最新淨值日期","淨值"]]
        法人級別_output = 法人級別_output.drop_duplicates()

        return 法人級別_output

